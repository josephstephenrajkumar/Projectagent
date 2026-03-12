"""
Tools: Excel Parser – comprehensive extraction of estimation-milestone
and ERP Excel files for the Project Creation flow.

Parses three sheets from the estimation-milestone Excel:
  1. Resources: month headers, TOTAL HOURS, per-resource breakdown, milestones, invoicing, revenue
  2. Travel & Expense: per-resource travel costs per month
  3. Other Costs: risk costs and other costs per month

Also parses ERP Excel files for Project table metadata.

Adapted from azureai_search_sow_crawler_agent_b_(_backup_).py
"""
import json
import re
from datetime import datetime
from typing import Optional


def _parse_date_header(val) -> Optional[str]:
    """Convert Excel date header to ISO date string."""
    if val is None:
        return None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%Y-%m-%d", "%d %b %Y", "%b %Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        return float(val)
    except (ValueError, TypeError):
        return default


def _parse_milestone_cell(cell_val) -> list[dict]:
    """Parse entries like 'PO Issue, 15/04/2024, Invoicing, SGD 34750.00'."""
    if not cell_val or not isinstance(cell_val, str):
        return []
    items = []
    for line in str(cell_val).strip().split("\n"):
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(",")]
        if len(parts) >= 4:
            amount_str = re.sub(r"[A-Z]{3}\s*", "", parts[3]).strip()
            items.append({
                "detail": parts[0],
                "date": parts[1],
                "type": parts[2],
                "amount": _safe_float(amount_str),
                "currency": re.search(r"[A-Z]{3}", parts[3]).group() if re.search(r"[A-Z]{3}", parts[3]) else "",
            })
        elif len(parts) >= 2:
            items.append({
                "detail": parts[0],
                "date": parts[1] if len(parts) > 1 else "",
                "type": parts[2] if len(parts) > 2 else "",
                "amount": 0.0,
                "currency": "",
            })
    return items


def _find_date_columns(ws, max_col):
    """Find date header columns and return {col_idx: iso_date} dict."""
    date_headers = {}
    for col_idx in range(1, max_col + 1):
        cell_val = ws.cell(row=1, column=col_idx).value
        parsed = _parse_date_header(cell_val)
        if parsed and re.match(r"\d{4}-\d{2}-\d{2}", parsed):
            date_headers[col_idx] = parsed
    return date_headers


def _find_header_columns(ws, max_col):
    """Map lowercase header names to column indices from row 1."""
    header_map = {}
    for col_idx in range(1, max_col + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val and isinstance(val, str):
            header_map[val.strip().lower()] = col_idx
    return header_map


def _find_row_by_label(ws, max_row, label_contains: str) -> Optional[int]:
    """Find the first row where column A contains the given text (case-insensitive)."""
    for row_idx in range(1, max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and isinstance(cell_val, str) and label_contains.lower() in cell_val.lower():
            return row_idx
    return None


def _extract_month_labels(ws, date_headers: dict) -> dict:
    """Extract month labels (1, 2, 3...) from the MONTH row."""
    month_row_idx = _find_row_by_label(ws, ws.max_row, "MONTH")
    month_labels = {}
    if month_row_idx:
        for col_idx, date_str in date_headers.items():
            val = ws.cell(row=month_row_idx, column=col_idx).value
            month_labels[date_str] = int(_safe_float(val)) if val is not None else None
    return month_labels


def parse_estimation_excel(file_path: str) -> dict:
    """
    Comprehensive parser for estimation-milestone Excel files.

    Returns dict with:
      - planned_months: [{month_number, date}]
      - startdateBaseline, endDateBaseline
      - total_cost, total_hours
      - total_hours_per_month: {date: hours}
      - resources: [{name, specialty, notes, monthly_hours, total_hours, cost_per_hour, total_cost, list_price, adjusted_rate, total_fees, billable}]
      - invoicing: [{detail, date, type, amount, currency, month_column}]
      - revenue: [{detail, date, type, amount, currency, month_column}]
      - travel_expenses: {total, per_resource: [{name, monthly_costs, total, billable_cost, non_billable_cost}]}
      - other_costs: {total, items: [{name, monthly_costs, total_costs, total_fees, billable}]}
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    result = {
        "planned_months": [],
        "startdateBaseline": None,
        "endDateBaseline": None,
        "total_cost": 0.0,
        "total_hours": 0,
        "total_fees": 0.0,
        "total_hours_per_month": {},
        "resources": [],
        "invoicing": [],
        "revenue": [],
        "travel_expenses": {"total": 0.0, "per_resource": []},
        "other_costs": {"total": 0.0, "items": []},
    }

    # ═══════════════════════════════════════════════════════════════════════════
    #  1. RESOURCES SHEET
    # ═══════════════════════════════════════════════════════════════════════════
    resources_ws = None
    for name in wb.sheetnames:
        if "resource" in name.lower():
            resources_ws = wb[name]
            break
    if resources_ws is None and len(wb.sheetnames) > 0:
        resources_ws = wb[wb.sheetnames[0]]

    if resources_ws:
        max_col = resources_ws.max_column
        max_row = resources_ws.max_row
        date_headers = _find_date_columns(resources_ws, max_col)
        header_map = _find_header_columns(resources_ws, max_col)

        # Baseline dates (first and last date column)
        if date_headers:
            sorted_dates = sorted(date_headers.values())
            result["startdateBaseline"] = sorted_dates[0]
            result["endDateBaseline"] = sorted_dates[-1]

        # Month labels (MONTH row: 1, 2, 3...)
        month_labels = _extract_month_labels(resources_ws, date_headers)
        result["planned_months"] = [
            {"month_number": month_labels.get(d), "date": d}
            for d in sorted(date_headers.values())
        ]

        # Column positions
        hours_col = header_map.get("hours")
        effort_col = header_map.get("effort needs")
        list_price_col = header_map.get("list price")
        adj_rate_col = header_map.get("adjusted rate")
        total_fees_col = header_map.get("total fees")
        cost_col = header_map.get("cost")
        total_cost_col = header_map.get("total cost")
        billable_col = header_map.get("billable")

        # Find key rows
        total_hours_row = _find_row_by_label(resources_ws, max_row, "TOTAL HOURS PER MONTH")
        milestone_row = _find_row_by_label(resources_ws, max_row, "MILESTONE")

        # --- Total Hours Per Month ---
        if total_hours_row and date_headers:
            for col_idx, date_str in date_headers.items():
                val = resources_ws.cell(row=total_hours_row, column=col_idx).value
                result["total_hours_per_month"][date_str] = _safe_float(val)

            if hours_col:
                result["total_hours"] = int(_safe_float(resources_ws.cell(row=total_hours_row, column=hours_col).value))
            if total_cost_col:
                result["total_cost"] = _safe_float(resources_ws.cell(row=total_hours_row, column=total_cost_col).value)
            if total_fees_col:
                result["total_fees"] = _safe_float(resources_ws.cell(row=total_hours_row, column=total_fees_col).value)

        # --- Per-Resource Breakdown (rows after TOTAL HOURS PER MONTH) ---
        if total_hours_row and date_headers:
            for row_idx in range(total_hours_row + 1, max_row + 1):
                resource_name = resources_ws.cell(row=row_idx, column=1).value
                if not resource_name or (isinstance(resource_name, str) and not resource_name.strip()):
                    break

                monthly_hours = {}
                for col_idx, date_str in date_headers.items():
                    monthly_hours[date_str] = _safe_float(resources_ws.cell(row=row_idx, column=col_idx).value)

                entry = {
                    "name": str(resource_name).strip(),
                    "specialty": str(resources_ws.cell(row=row_idx, column=2).value or "").strip(),
                    "notes": str(resources_ws.cell(row=row_idx, column=3).value or "").strip(),
                    "monthly_hours": monthly_hours,
                    "total_hours": int(_safe_float(resources_ws.cell(row=row_idx, column=hours_col).value)) if hours_col else 0,
                    "effort_needs": _safe_float(resources_ws.cell(row=row_idx, column=effort_col).value) if effort_col else 0,
                    "list_price": _safe_float(resources_ws.cell(row=row_idx, column=list_price_col).value) if list_price_col else 0,
                    "adjusted_rate": _safe_float(resources_ws.cell(row=row_idx, column=adj_rate_col).value) if adj_rate_col else 0,
                    "total_fees": _safe_float(resources_ws.cell(row=row_idx, column=total_fees_col).value) if total_fees_col else 0,
                    "cost_per_hour": _safe_float(resources_ws.cell(row=row_idx, column=cost_col).value) if cost_col else 0,
                    "total_cost": _safe_float(resources_ws.cell(row=row_idx, column=total_cost_col).value) if total_cost_col else 0,
                    "billable": str(resources_ws.cell(row=row_idx, column=billable_col).value or "").strip() if billable_col else "",
                }
                result["resources"].append(entry)

        # --- Milestones, Invoicing & Revenue (rows 4, 5, 6 typically) ---
        if milestone_row and date_headers:
            # Scan milestone row + next few rows until EFFORT NEEDS or empty
            effort_row = _find_row_by_label(resources_ws, max_row, "EFFORT NEEDS")
            scan_end = effort_row if effort_row else milestone_row + 4

            for row_idx in range(milestone_row, scan_end):
                if row_idx > max_row:
                    break
                for col_idx, date_str in date_headers.items():
                    cell_val = resources_ws.cell(row=row_idx, column=col_idx).value
                    if cell_val and isinstance(cell_val, str) and cell_val.strip():
                        parsed_items = _parse_milestone_cell(cell_val)
                        for item in parsed_items:
                            item["month_column"] = date_str
                            item_type = item.get("type", "").lower()
                            if "invoic" in item_type:
                                result["invoicing"].append(item)
                            elif "revenue" in item_type:
                                result["revenue"].append(item)
                            else:
                                # Default: invoicing-like items (PO, P1, etc.)
                                result["invoicing"].append(item)

    # ═══════════════════════════════════════════════════════════════════════════
    #  2. TRAVEL & EXPENSE SHEET
    # ═══════════════════════════════════════════════════════════════════════════
    travel_ws = None
    for name in wb.sheetnames:
        if "travel" in name.lower():
            travel_ws = wb[name]
            break

    if travel_ws:
        t_date_headers = _find_date_columns(travel_ws, travel_ws.max_column)
        t_header_map = _find_header_columns(travel_ws, travel_ws.max_column)
        t_total_col = t_header_map.get("total")
        t_billable_cost_col = t_header_map.get("billable cost")
        t_non_billable_col = t_header_map.get("non billable cost")
        t_billable_col = t_header_map.get("billable")

        # Find "COST PER MONTH" row
        cost_per_month_row = _find_row_by_label(travel_ws, travel_ws.max_row, "COST PER MONTH")

        if cost_per_month_row:
            # Total travel cost from the COST PER MONTH summary row
            if t_total_col:
                result["travel_expenses"]["total"] = _safe_float(
                    travel_ws.cell(row=cost_per_month_row, column=t_total_col).value
                )

            # Per-resource travel costs (rows after COST PER MONTH)
            for row_idx in range(cost_per_month_row + 1, travel_ws.max_row + 1):
                res_name = travel_ws.cell(row=row_idx, column=1).value
                if not res_name or (isinstance(res_name, str) and not res_name.strip()):
                    break

                monthly_costs = {}
                for col_idx, date_str in t_date_headers.items():
                    monthly_costs[date_str] = _safe_float(travel_ws.cell(row=row_idx, column=col_idx).value)

                entry = {
                    "name": str(res_name).strip(),
                    "notes": str(travel_ws.cell(row=row_idx, column=3).value or "").strip(),
                    "monthly_costs": monthly_costs,
                    "total": _safe_float(travel_ws.cell(row=row_idx, column=t_total_col).value) if t_total_col else 0,
                    "billable_cost": _safe_float(travel_ws.cell(row=row_idx, column=t_billable_cost_col).value) if t_billable_cost_col else 0,
                    "non_billable_cost": _safe_float(travel_ws.cell(row=row_idx, column=t_non_billable_col).value) if t_non_billable_col else 0,
                    "billable": str(travel_ws.cell(row=row_idx, column=t_billable_col).value or "").strip() if t_billable_col else "",
                }
                result["travel_expenses"]["per_resource"].append(entry)

    # ═══════════════════════════════════════════════════════════════════════════
    #  3. OTHER COSTS SHEET
    # ═══════════════════════════════════════════════════════════════════════════
    other_ws = None
    for name in wb.sheetnames:
        if "other" in name.lower():
            other_ws = wb[name]
            break

    if other_ws:
        o_date_headers = _find_date_columns(other_ws, other_ws.max_column)
        o_header_map = _find_header_columns(other_ws, other_ws.max_column)
        o_total_costs_col = o_header_map.get("total costs")
        o_total_fees_col = o_header_map.get("total fees")
        o_billable_col = o_header_map.get("billable")

        cost_per_month_row = _find_row_by_label(other_ws, other_ws.max_row, "COST PER MONTH")

        if cost_per_month_row:
            # Total other costs from summary row
            if o_total_costs_col:
                result["other_costs"]["total"] = _safe_float(
                    other_ws.cell(row=cost_per_month_row, column=o_total_costs_col).value
                )

            # Individual cost items (rows after COST PER MONTH, e.g. "Risk Cost")
            for row_idx in range(cost_per_month_row + 1, other_ws.max_row + 1):
                cost_name = other_ws.cell(row=row_idx, column=1).value
                if not cost_name or (isinstance(cost_name, str) and not cost_name.strip()):
                    break

                monthly_costs = {}
                for col_idx, date_str in o_date_headers.items():
                    monthly_costs[date_str] = _safe_float(other_ws.cell(row=row_idx, column=col_idx).value)

                entry = {
                    "name": str(cost_name).strip(),
                    "monthly_costs": monthly_costs,
                    "total_costs": _safe_float(other_ws.cell(row=row_idx, column=o_total_costs_col).value) if o_total_costs_col else 0,
                    "total_fees": _safe_float(other_ws.cell(row=row_idx, column=o_total_fees_col).value) if o_total_fees_col else 0,
                    "billable": str(other_ws.cell(row=row_idx, column=o_billable_col).value or "").strip() if o_billable_col else "",
                }
                result["other_costs"]["items"].append(entry)

    wb.close()
    return result


def parse_erp_excel(file_path: str) -> dict:
    """
    Parse an ERP Project Data Excel file and return project metadata.
    Looks for common ERP field names in the first sheet
    (key-value pairs or table rows).
    """
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    erp_data = {}

    # Strategy 1: Key-Value pairs (col A = field name, col B = value)
    for row_idx in range(1, ws.max_row + 1):
        key = ws.cell(row=row_idx, column=1).value
        val = ws.cell(row=row_idx, column=2).value
        if key and isinstance(key, str) and val is not None:
            erp_data[key.strip()] = val

    # Strategy 2: If first row looks like headers, treat as a table
    if ws.max_row >= 2:
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            h = ws.cell(row=1, column=col_idx).value
            if h and isinstance(h, str):
                headers.append((col_idx, h.strip()))

        if headers and ws.max_row >= 2:
            # Take the first data row (row 2)
            for col_idx, header in headers:
                val = ws.cell(row=2, column=col_idx).value
                if val is not None:
                    erp_data[header] = val

    wb.close()
    return erp_data


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        data = parse_estimation_excel(sys.argv[1])
        print(json.dumps(data, indent=2, default=str))
    else:
        print("Usage: python excel_parser.py <path_to_xlsx>")
